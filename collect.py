# -*- coding: utf-8 -*-
"""코스닥 상장폐지 리스크 모니터링 데이터 수집기.

시세 소스 (둘 중 자동 선택):
  1) pykrx (KRX)  — KRX_ID/KRX_PW 환경변수 설정 시. 시가총액 이력까지 정확.
     ※ KRX 정보데이터시스템이 2025년부터 로그인(무료 계정)을 요구함.
  2) 네이버 금융  — 인증 불필요 fallback. 시총 이력은 (현재시총×종가비율) 근사.

재무 소스: OpenDART 사업보고서 (DART_API_KEY 설정 시) — 매출·영업이익·부채비율·자본잠식

출력: docs/data.json
캐시: dart_cache.json (확정된 연간 재무는 재요청하지 않음)

환경변수:
  DART_API_KEY   : OpenDART 인증키 (없으면 재무 항목은 비운 채 진행)
  KRX_ID, KRX_PW : KRX 정보데이터시스템 계정 (없으면 네이버 소스 사용)
  TICKER_LIMIT   : 테스트용 — 처리할 종목 수 제한 (예: 10)
"""
import datetime as dt
import io
import json
import os
import re
import sys
import time
import zipfile
import xml.etree.ElementTree as ET

import requests

MCAP_LIMIT = 300  # 억 원 — 스크리닝 기준 ('27.1월 시행 기준 선반영)
MCAP_RULE_NOW = 200  # 억 원 — 현행 관리종목 지정 기준
PENNY_PRICE = 1000  # 원
LOOKBACK_TRADING_DAYS = 130  # 연속일수 계산용 시세 조회 기간(거래일)
FIN_YEARS = [2023, 2024, 2025]

DART_KEY = os.environ.get("DART_API_KEY", "").strip()
USE_KRX = bool(os.environ.get("KRX_ID") and os.environ.get("KRX_PW"))
TICKER_LIMIT = int(os.environ.get("TICKER_LIMIT", "0") or 0)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_PATH = os.path.join(BASE_DIR, "dart_cache.json")
OUT_PATH = os.path.join(BASE_DIR, "docs", "data.json")
XLSX_PATH = os.path.join(BASE_DIR, "docs", "data.xlsx")
PROFILE_PATH = os.path.join(BASE_DIR, "profile_cache.json")

EOK = 100_000_000  # 1억
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}


def log(msg):
    print(msg, flush=True)


def streak(values, pred):
    """뒤에서부터 pred를 만족하는 연속 개수."""
    n = 0
    for v in reversed(values):
        if pred(v):
            n += 1
        else:
            break
    return n


# ---------------------------------------------------------------- 시세: pykrx (KRX)

def universe_krx():
    from pykrx import stock
    d = dt.date.today()
    for _ in range(10):
        ymd = d.strftime("%Y%m%d")
        try:
            df = stock.get_market_cap_by_ticker(ymd, market="KOSDAQ")
        except Exception:
            df = None
        if df is not None and not df.empty and int(df["시가총액"].sum()) > 0:
            rows = []
            for t, row in df.iterrows():
                mcap = row["시가총액"] / EOK
                rows.append({
                    "code": t,
                    # 이름 조회는 느리므로 스크리닝 대상만
                    "name": stock.get_market_ticker_name(t) if 0 < mcap < MCAP_LIMIT else "",
                    "close": int(row["종가"]),
                    "mcap": round(mcap, 1),
                })
            return ymd, rows
        d -= dt.timedelta(days=1)
    raise RuntimeError("최근 10일 내 코스닥 시세를 찾지 못했습니다")


def history_krx(code, base_ymd):
    """(종가 리스트, 시총(억) 리스트) — 과거→현재 순."""
    from pykrx import stock
    frm = (dt.datetime.strptime(base_ymd, "%Y%m%d")
           - dt.timedelta(days=LOOKBACK_TRADING_DAYS * 2)).strftime("%Y%m%d")
    df = stock.get_market_cap(frm, base_ymd, code)
    if df is None or df.empty:
        return [], []
    mcaps = [m / EOK for m in df["시가총액"].tolist()]
    if "종가" in df.columns:
        closes = df["종가"].tolist()
    else:
        # 기간 조회에는 종가 컬럼이 없음 — 시가총액/상장주식수로 복원
        closes = [(int(m / s) if s else 0)
                  for m, s in zip(df["시가총액"].tolist(), df["상장주식수"].tolist())]
    return closes, mcaps


# ---------------------------------------------------------------- 시세: 네이버 금융

def _num(s):
    s = str(s).replace(",", "").strip()
    return float(s) if s and s != "-" else 0.0


def universe_naver():
    """코스닥 전 종목 (시총 억원). marketValue 오름차순 아님 — 전체 페이지 수집."""
    rows, page = [], 1
    while True:
        r = requests.get(
            f"https://m.stock.naver.com/api/stocks/marketValue/KOSDAQ",
            params={"page": page, "pageSize": 100}, headers=UA, timeout=30)
        r.raise_for_status()
        data = r.json()
        for s in data.get("stocks", []):
            rows.append({
                "code": s["itemCode"],
                "name": s["stockName"],
                "close": int(_num(s["closePrice"])),
                "mcap": _num(s["marketValue"]),  # 이미 억 원 단위
            })
        total = data.get("totalCount", 0)
        if page * 100 >= total or not data.get("stocks"):
            break
        page += 1
        time.sleep(0.1)
    base_ymd = dt.date.today().strftime("%Y%m%d")
    return base_ymd, rows


def history_naver(code, cur_close, cur_mcap):
    """fchart 일봉으로 종가 이력 조회. 시총 이력은 종가 비율로 근사."""
    r = requests.get(
        "https://fchart.stock.naver.com/sise.nhn",
        params={"symbol": code, "timeframe": "day",
                "count": LOOKBACK_TRADING_DAYS, "requestType": 0},
        headers=UA, timeout=30)
    r.raise_for_status()
    closes, last_ymd = [], None
    for m in re.finditer(r'data="([^"]+)"', r.text):
        parts = m.group(1).split("|")
        if len(parts) >= 5:
            last_ymd = parts[0]
            closes.append(int(_num(parts[4])))
    if not closes or cur_close <= 0:
        return [], [], last_ymd
    mcaps = [cur_mcap * c / cur_close for c in closes]
    return closes, mcaps, last_ymd


# ---------------------------------------------------------------- 기업 프로필 (네이버)

def naver_profile(code):
    """업종명 + 사업내용 한 줄 (finance.naver 기업개요, FnGuide 제공)."""
    r = requests.get("https://finance.naver.com/item/main.naver",
                     params={"code": code}, headers=UA, timeout=20)
    r.encoding = "utf-8"
    html = r.text
    m = re.search(r"upjong[^>]*>([^<]+)</a>", html)
    sector = m.group(1).strip() if m else None
    biz = None
    s = re.search(r'summary_info[^>]*>(.*?)</div>', html, re.S)
    if s:
        txt = " ".join(re.sub(r"<[^>]+>", " ", s.group(1)).split())
        txt = txt.replace("기업개요", "", 1).strip()
        sents = [x.strip() for x in re.split(r"(?<=[음됨임함])\.\s*", txt) if x.strip()]
        # 설립·상장 연혁 문장은 건너뛰고 사업을 설명하는 문장 선택
        biz = next((x for x in sents if not re.search(r"설립|상장", x)),
                   sents[0] if sents else None)
        if biz and len(biz) > 90:
            biz = biz[:90] + "…"
    return {"sector": sector, "biz": biz}


def naver_flags(code):
    """관리종목 지정·거래정지 여부 (매일 갱신)."""
    r = requests.get(f"https://m.stock.naver.com/api/stock/{code}/basic",
                     headers=UA, timeout=15)
    r.raise_for_status()
    d = r.json()
    halted = ((d.get("tradeStopType") or {}).get("name") == "HALTED")
    return bool(d.get("isManagement")), halted


# ---------------------------------------------------------------- OpenDART

def dart_corp_map():
    """종목코드 → DART corp_code 매핑."""
    r = requests.get(
        "https://opendart.fss.or.kr/api/corpCode.xml",
        params={"crtfc_key": DART_KEY}, timeout=60)
    r.raise_for_status()
    zf = zipfile.ZipFile(io.BytesIO(r.content))
    root = ET.fromstring(zf.read(zf.namelist()[0]))
    mapping = {}
    for el in root.iter("list"):
        stock_code = (el.findtext("stock_code") or "").strip()
        if stock_code:
            mapping[stock_code] = el.findtext("corp_code").strip()
    return mapping


def dart_fin_year(corp_code, year):
    """사업보고서 주요계정. 연결(CFS) 우선, 없으면 별도(OFS)."""
    r = requests.get(
        "https://opendart.fss.or.kr/api/fnlttSinglAcnt.json",
        params={"crtfc_key": DART_KEY, "corp_code": corp_code,
                "bsns_year": str(year), "reprt_code": "11011"},
        timeout=30)
    r.raise_for_status()
    data = r.json()
    if data.get("status") != "000":
        return None

    def amount(txt):
        txt = (txt or "").replace(",", "").strip()
        if not txt or txt == "-":
            return None
        try:
            return round(int(txt) / EOK, 1)
        except ValueError:
            return None

    by_fs = {"CFS": {}, "OFS": {}}
    for item in data.get("list", []):
        acc = by_fs.get(item.get("fs_div"))
        if acc is not None:
            acc[item.get("account_nm", "").strip()] = amount(item.get("thstrm_amount"))

    for fs in ("CFS", "OFS"):
        acc = by_fs[fs]
        if not acc:
            continue
        if acc.get("매출액") is None and acc.get("영업이익") is None \
                and acc.get("자산총계") is None:
            continue
        return {
            "rev": acc.get("매출액"),
            "op": acc.get("영업이익"),
            "ni": acc.get("당기순이익"),
            "assets": acc.get("자산총계"),
            "liab": acc.get("부채총계"),
            "equity": acc.get("자본총계"),
            "fs": fs,
        }
    return None


# ---------------------------------------------------------------- 엑셀

def build_xlsx(out):
    """docs/data.xlsx 생성 — 웹의 '엑셀 다운로드' 버튼용."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "시총300억미만"

    years = [str(y) for y in out["fin_years"]]
    headers = (["번호", "기업명", "종목코드", "업종", "사업내용",
                "주가(원)", "시총(억)", "부채비율(%)", "자산(억)", "부채(억)", "자본(억)"]
               + [f"매출 {y[2:]}" for y in years]
               + [f"영업이익 {y[2:]}" for y in years]
               + ["동전주 연속(일)", "시총200억미달 연속(일)", "관리종목", "거래정지", "자본잠식"])

    head_fill = PatternFill("solid", fgColor="C6E0B4")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    fin_fmt = "#,##0.0;[Red](#,##0.0)"
    fin_first, fin_last = 12, 11 + 2 * len(years)  # 매출·영업이익 열 범위
    for r, comp in enumerate(out["companies"], 2):
        fin = comp.get("fin", {})
        row = ([r - 1, comp["name"], comp["code"],
                comp.get("sector"), comp.get("biz"), comp["close"],
                comp["mcap"], comp["debt_ratio"],
                comp.get("assets"), comp.get("liab"), comp.get("equity")]
               + [fin.get(y, {}).get("rev") for y in years]
               + [fin.get(y, {}).get("op") for y in years]
               + [comp["penny_streak"] or None,
                  comp["under200_streak"] or None,
                  "관리" if comp.get("is_management") else None,
                  "정지" if comp.get("trade_stop") else None,
                  "잠식" if comp["equity_impaired"] else None])
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.font = Font(size=10)
            if c == 6:
                cell.number_format = "#,##0"
            elif c in (7, 8):
                cell.number_format = "#,##0.0"
            elif 9 <= c <= 11 or fin_first <= c <= fin_last:
                cell.number_format = fin_fmt
            if c in (1, 3, 4) or c > fin_last:
                cell.alignment = center

    widths = ([6, 16, 9, 15, 60, 9, 8, 10, 9, 9, 9]
              + [10] * (2 * len(years)) + [12, 16, 8, 8, 8])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(out['companies']) + 1}"

    wb.save(XLSX_PATH)
    log(f"엑셀 생성: {XLSX_PATH}")


# ---------------------------------------------------------------- main

def main():
    source = "KRX(pykrx)" if USE_KRX else "네이버 금융(fallback)"
    log(f"시세 소스: {source}")

    if USE_KRX:
        base_ymd, rows = universe_krx()
    else:
        base_ymd, rows = universe_naver()
    log(f"코스닥 {len(rows)}종목 수집")

    # 보통주만(종목코드 끝자리 0) — 우선주 제외, SPAC('기업인수목적'/'스팩') 제외
    under = [r for r in rows if 0 < r["mcap"] < MCAP_LIMIT
             and r["code"].endswith("0")
             and "기업인수목적" not in r["name"]
             and "스팩" not in r["name"]]
    under.sort(key=lambda r: r["mcap"])
    log(f"시총 {MCAP_LIMIT}억 미만 (SPAC 제외): {len(under)}종목")

    if TICKER_LIMIT:
        under = under[:TICKER_LIMIT]
        log(f"TICKER_LIMIT={TICKER_LIMIT} 적용")

    cache = {}
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, encoding="utf-8") as f:
            cache = json.load(f)

    profiles = {}
    if os.path.exists(PROFILE_PATH):
        with open(PROFILE_PATH, encoding="utf-8") as f:
            profiles = json.load(f)

    corp_map = {}
    if DART_KEY:
        try:
            corp_map = dart_corp_map()
            log(f"DART corp_code {len(corp_map)}건 로드")
        except Exception as e:
            log(f"[경고] DART corp_code 로드 실패: {e}")
    else:
        log("[경고] DART_API_KEY 미설정 — 재무 항목 없이 진행")

    companies = []
    for i, r in enumerate(under, 1):
        code = r["code"]
        try:
            if USE_KRX:
                closes, mcaps = history_krx(code, base_ymd)
            else:
                closes, mcaps, last_ymd = history_naver(code, r["close"], r["mcap"])
                if last_ymd:
                    base_ymd = max(base_ymd, last_ymd) if i == 1 else base_ymd
        except Exception as e:
            log(f"[경고] {code} 시세 이력 실패: {e}")
            closes, mcaps = [], []
        time.sleep(0.15)

        penny = streak(closes, lambda c: 0 < c < PENNY_PRICE)
        under300 = streak(mcaps, lambda m: 0 < m < MCAP_LIMIT)
        under200 = streak(mcaps, lambda m: 0 < m < MCAP_RULE_NOW)

        # 업종·사업내용은 캐시(변동 거의 없음), 관리종목·거래정지는 매일 조회
        prof = profiles.get(code)
        if not prof or not prof.get("sector"):
            try:
                prof = naver_profile(code)
                profiles[code] = prof
            except Exception as e:
                log(f"[경고] {code} 프로필 실패: {e}")
                prof = prof or {"sector": None, "biz": None}
            time.sleep(0.1)
        is_mgmt = halted = False
        try:
            is_mgmt, halted = naver_flags(code)
        except Exception as e:
            log(f"[경고] {code} 관리종목 플래그 실패: {e}")
        time.sleep(0.08)

        fin = {}
        debt_ratio = None
        equity_impaired = None
        assets = liab = equity = None
        corp_code = corp_map.get(code)
        if corp_code:
            for year in FIN_YEARS:
                key = f"{corp_code}:{year}"
                if key in cache:
                    fy = cache[key]
                else:
                    try:
                        fy = dart_fin_year(corp_code, year)
                    except Exception as e:
                        log(f"[경고] {code}/{year} DART 실패: {e}")
                        fy = None
                    # 과거 연도는 빈 응답도 캐시, 최신 연도 미공시만 다음 실행에서 재시도
                    if fy is not None or year < max(FIN_YEARS):
                        cache[key] = fy
                    time.sleep(0.12)
                if fy:
                    fin[str(year)] = fy
            latest = next((fin[str(y)] for y in sorted(FIN_YEARS, reverse=True)
                           if str(y) in fin), None)
            if latest:
                assets = latest.get("assets")
                liab = latest.get("liab")
                equity = latest.get("equity")
                if equity is not None:
                    equity_impaired = equity <= 0
                    if equity > 0 and liab is not None:
                        debt_ratio = round(liab / equity * 100, 1)

        companies.append({
            "code": code,
            "name": r["name"],
            "sector": prof.get("sector"),
            "biz": prof.get("biz"),
            "close": r["close"],
            "mcap": round(r["mcap"], 1),
            "penny_streak": penny,
            "under300_streak": under300,
            "under200_streak": under200,
            "debt_ratio": debt_ratio,
            "equity_impaired": equity_impaired,
            "assets": assets,
            "liab": liab,
            "equity": equity,
            "is_management": is_mgmt,
            "trade_stop": halted,
            "fin": fin,
        })
        if i % 25 == 0:
            log(f"  {i}/{len(under)} 처리")
            # 중간 캐시 저장 — 중단돼도 수집분은 보존
            with open(CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False)
            with open(PROFILE_PATH, "w", encoding="utf-8") as f:
                json.dump(profiles, f, ensure_ascii=False)

    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    with open(PROFILE_PATH, "w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False)

    out = {
        "updated_kst": (dt.datetime.now(dt.timezone.utc)
                        + dt.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M"),
        "base_date": base_ymd,
        "source": source,
        "mcap_limit": MCAP_LIMIT,
        "fin_years": FIN_YEARS,
        "has_dart": bool(corp_map),
        "companies": companies,
    }
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)
    build_xlsx(out)
    log(f"완료: {OUT_PATH} ({len(companies)}종목)")


if __name__ == "__main__":
    if "--xlsx-only" in sys.argv:
        # 기존 data.json으로 엑셀만 재생성 (재수집 없이)
        with open(OUT_PATH, encoding="utf-8") as f:
            build_xlsx(json.load(f))
        sys.exit(0)
    sys.exit(main())
